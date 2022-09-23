from openpyxl import load_workbook

# 1.在原文件写入
wb = load_workbook('file/p2.xlsx')
sheet = wb.worksheets[0]
# 获取内容
sheet.cell(1, 1).value = '新的开始'  # 改变
wb.save('file/p1.xlsx')  # 保存  文件必须关闭才不会报错

# 2.新创建Excel
from openpyxl import workbook

# 创建一个新的Excel，并创建一个sheet，名字就叫sheet
wb = workbook.Workbook()
sheet = wb.worksheets[0]
sheet.cell(1, 1).value = '新的开始'
wb.save('file/p2.xlsx')  # 也要保存

# ##针对sheet
# 3.修改sheet名称   sheet.title=''
wb = workbook.Workbook()
sheet = wb.worksheets[0]
sheet.title = '数据集'
wb.save('file/p3.xlsx')

# 4.创建新的sheet，并设置颜色   wb.create_sheet(名字，位置)
sheet = wb.create_sheet('工作计划', 1)
sheet.sheet_properties.tabColor = '66CDAA'  # .sheet_properties.tabColor='RGB颜色对照表'
wb.save('file/p3.xlsx')

# 5.默认打开的sheet   .active=0
wb.active = 1
wb.save('file/p3.xlsx')

# 6.拷贝sheet
new_sheet = wb.copy_worksheet(wb.worksheets[1])
new_sheet.title = '成绩'
wb.save('file/p3.xlsx')

# 7.删除sheet   del
del wb['数据集']
wb.save('file/p3.xlsx')

# ##针对单元格
# 8.获取某个单元格，修改值 .value
wb = load_workbook('file/p3.xlsx')
sheet = wb.worksheets[0]
cell = sheet.cell(1, 1)
cell.value = '开始'
wb.save('file/p3.xlsx')

# 9.索引修改值
sheet['B3'] = 'hello'
wb.save('file/p3.xlsx')

# 10.获取某些单元格，修改值
wb = load_workbook('file/p4.xlsx')
sheet = wb.worksheets[0]
print(sheet['B11':'C12'])  # ((<Cell '数据导出'.B2>, <Cell '数据导出'.C2>), (<Cell '数据导出'.B3>, <Cell '数据导出'.C3>))
for row in sheet['B11':'C12']:  # 循环每一行
    for cell in row:  # 每一行的每个单元格
        cell.value = '新的值'  # 赋予新值
wb.save('file/p5.xlsx')

# 11.改变对齐方式
# horizontal，水平方向对齐方式："general", "left", "center", "right", "fill", "justify", "centerContinuous", "distributed"
# vertical，垂直方向对齐方式："top", "center", "bottom", "justify", "distributed"
# text_rotation，旋转角度。
# wrap_text，是否自动换行。
cell = sheet.cell(1, 1)
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, GradientFill

cell.alignment = Alignment(horizontal='center', vertical='distributed', text_rotation=45, wrap_text=True)
wb.save('file/p5.xlsx')

# 12.边框
# side的style有如下：dashDot','dashDotDot', 'dashed','dotted','double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot','mediumDashed', 'slantDashDot', 'thick', 'thin'
cell = sheet.cell(9, 2)
cell.border = Border(
    top=Side(style="thin", color="FFB6C1"),
    bottom=Side(style="dashed", color="FFB6C1"),
    left=Side(style="dashed", color="FFB6C1"),
    right=Side(style="dashed", color="9932CC"),
    diagonal=Side(style="thin", color="483D8B"),  # 对角线
    diagonalUp=True,  # 左下 ~ 右上
    diagonalDown=True  # 左上 ~ 右下
)
wb.save("file/p4.xlsx")

# 13.字体   Font
cell = sheet.cell(5, 1)
cell.font = Font(name='黑体', size=45, color='ff0000', underline='single')  # 下划线
wb.save('file/p4.xlsx')

# 14.背景色   PatternFill
cell = sheet.cell(2, 3)
cell.fill = PatternFill('solid', fgColor='ff0000')
wb.save('file/p4.xlsx')

# 15.渐变背景色   GradientFill
cell = sheet.cell(5, 5)
cell.fill = GradientFill("linear", stop=("FFFFFF", "99ccff", "000000"))
wb.save("file/p2.xlsx")

# 16.设置宽和高
sheet.row_dimensions[1].height = 50  # 第一行
sheet.column_dimensions["E"].width = 100  # E列
wb.save("file/p4.xlsx")

# 17.合并单元格   .merge_cells
sheet.merge_cells("B2:D8")
sheet.merge_cells(start_row=15, start_column=3, end_row=18, end_column=8)
wb.save("file\p4.xlsx")
# 解除合并
sheet.unmerge_cells("B2:D8")
wb.save("file/p4.xlsx")

# 18.写入公式
sheet = wb.worksheets[3]
sheet["D1"] = "合计"
sheet["D2"] = "=B2*C2"
wb.save("file/p4.xlsx")

sheet = wb.worksheets[3]
sheet["D3"] = "=SUM(B3,C3)"  # 也可以写Excel内部函数
wb.save("file/p4.xlsx")

# 19.删除   .delete_rows()
wb = load_workbook('file/p5.xlsx')
sheet = wb.worksheets[0]
sheet.delete_rows(idx=20, amount=2)  # 从idx开始删，往后删两行
sheet.delete_cols(idx=20, amount=2)
wb.save('file/p5.xlsx')

# 20.插入
sheet.insert_rows(idx=5, amount=10)
sheet.insert_cols(idx=3, amount=2)
wb.save("file/p2.xlsx")

# 21.循环写内容
sheet = wb["Sheet"]
cell_range = sheet['A1:C2']
for row in cell_range:
    for cell in row:
        cell.value = "xx"
for row in sheet.iter_rows(min_row=5, min_col=1, max_col=7, max_row=10):
    for cell in row:
        cell.value = "oo"
wb.save("file/p2.xlsx")

# 22.移动
# 将H2:J10范围的数据，向右移动15个位置、向上移动1个位置
sheet.move_range("H2:J10", rows=1, cols=15)
wb.save("p2.xlsx")
sheet = wb.worksheets[3]
sheet["D1"] = "合计"
sheet["D2"] = "=B2*C2"
sheet["D3"] = "=SUM(B3,C3)"
sheet.move_range("B1:D3", cols=10, translate=True)  # 自动翻译公式
wb.save("p2.xlsx")

# 23.打印区域
sheet.print_area = "A1:D200"
wb.save("p2.xlsx")

# 24.打印时，每个页面的固定表头
sheet.print_title_cols = "A:D"
sheet.print_title_rows = "1:3"
wb.save("p2.xlsx")
