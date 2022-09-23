# 下载第三方模块openpyxl

# 一.读取Excel的sheet
from openpyxl import load_workbook

wk = load_workbook('file/p2.xlsx')

# 1.获取Excel文件中所有的sheet名称   .sheetnames
print(wk.sheetnames)

# 2.选择sheet ,基于sheet名称  .cell()
sheet = wk['数据导出']
cell = sheet.cell(1, 2)  # 读取第一行，第一列   必须写.cell(x,y)
print(cell.value)  # 打印值  .value

# 3.选择sheet，基于索引位置  .worksheets[x]
sheet = wk.worksheets[0]
print(sheet.cell(1, 3).value)

# 4.循环所有的sheet
for name in wk.sheetnames:
    sheet = wk[name]  # 2.
    print(sheet.cell(1, 1).value)

for sheet in wk.worksheets:  # 遍历索引，直接得到sheet
    print(sheet.cell(1, 1).value)

for sheet in wk:  # 也可以直接遍历文件，和上面结果一样
    print(sheet.cell(1, 1).value)

# 二.读单元格
wb = load_workbook('file/p2.xlsx')  # 读取文件
sheet = wb.worksheets[0]  # 选择sheet
# 1.获取单元格信息
cell = sheet.cell(1, 1)  # 单元格的位置是从1开始的，而不是0

print(cell.value)  # .value   单元格的文本信息

print(cell.style)  # .style   样式

print(cell.font)  # .font    字体

print(cell.alignment)  # .aligment   对齐信息

# 2.获取某个单元格   sheet['A2']
print(sheet['A2'].value)
print(sheet['D3'].value)

# 3.获取第n行所有的单元格   sheet[1]
print(sheet[1])
for cell in sheet[1]:
    print(cell.value)

# 4.获取所有行的数据 (获取某一列)   sheet.rows
print(sheet.rows)  # 获取所有行
for row in sheet.rows:
    # print(row)
    print(row[0].value)  # 获取所有行第一个单元格的数据

# 5.获取所有列的数据   sheet.columns
print(sheet.columns)
for col in sheet.columns:
    print(col[0].value)  # 获取所有列第一个单元格的内容

# ##三、读取合并单元格的内容
wb = load_workbook('file/p2.xlsx')
sheet = wb.worksheets[2]  # 第三个表格
print(sheet.cell(1, 1))  # <Cell 'Sheet1'.A1>
print(sheet.cell(1, 2).value)  # <MergedCell 'Sheet1'.B1>   合并的单元格,且内容为空
for row in sheet.rows:
    print(row)

