# 1.基于csv格式实现 用户的注册 & 登录认证。详细需求如下：
# - 用户注册时，新注册用户要写入文件csv文件中，输入Q或q则退出。
# - 用户登录时，逐行读取csv文件中的用户信息并进行校验。
# - 提示：文件路径须使用os模块构造的绝对路径的方式。
import os

'''
path = os.path.dirname(os.path.abspath(__file__))  # 获得上级目录
file_path = os.path.join(path, 'info.csv')  # 拼接目录
if not os.path.exists(file_path):  # 判断目录是否存在
    os.makedirs(file_path)  # 不存在就创建该目录
# 注册功能
while True:
    request = input('是否要进行注册？y/n')
    request = request.upper()
    if request not in {'Y', 'N'}:
        print('输入错误')
    if request == 'N':
        break
    with open(file_path, mode='a', encoding='utf-8') as file:
        while True:
            num = input('请输入注册的账号(q退出)：')
            if num == 'q':
                break
            pwd = input('请输入注册账号的密码：')
            file.write(f'账号：{num},密码：{pwd}\n')
            file.flush()
            print('注册成功')
    break
# 用户登录
print('欢迎使用登录系统')
if not os.path.exists(file_path):  # 判断文件是否存在
    print('用户文件不存在')
else:
    while True:
        user_name = input('请输入账号：')
        user_pwd = input('请输入密码：')
        with open(file_path, mode='r', encoding='utf-8') as file:
            for i in file:  # for ...else 语句
                if user_name == i.strip().split('：')[1][0:5] and user_pwd == i.strip().split('：')[2]:
                    print('登录成功')
                    break
            else:
                print('账号或密码错误，请重新输入')
                continue
            break
'''
'''
# 2.补充代码：实现去网上获取指定地区的天气信息，并写入到Excel中。
# -提取XML格式中的数据
# -为每个城市创建一个sheet，并将获取的xml格式中的数据写入到excel中。
import requests
from xml.etree import ElementTree as ET
from openpyxl import workbook

wb = workbook.Workbook()  # 创建一个新的excel
del wb['Sheet']
path = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(path, '天气.xlsx')
while True:
    city = input('请输入查询天气的城市（q或Q退出）：')
    if city == 'q' or city == 'Q':
        break
    url = f'http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={city}'
    requ = requests.get(url)
    # print(requ.text)  获取到了字符串
    root = ET.XML(requ.text)
    # print(root)
    sheet = wb.create_sheet(city)  # 为每个城市创建一个sheet
    for row, node in enumerate(root, 1):  # for  row, node in enumerate(root, 1)   row=1,2,3,4...  node就是root的值
        # print(row, node.text)
        sheet.cell(row, 1).value = node.text
wb.save(excel_path)
'''
'''
# 3.读取ini文件内容，按照规则写入到Excel中。
# -读取ini格式的文件，并创建一个excel文件，且为每个节点创建一个sheet，然后将节点下的键值写入到excel中，按照如下格式。
# -首行，字体白色 & 单元格背景色蓝色。
# -内容均居中。
# -边框。
import configparser
from openpyxl import workbook
from openpyxl.styles import Font, Fill, Alignment, Border, Side, PatternFill

wb = workbook.Workbook()  # 创建一个新的Excel文件
del wb['Sheet']  # 删除原来的sheet

path = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(path, 'my.ini')  # 获取文件路径

config = configparser.ConfigParser()
config.read(file_path, encoding='utf-8')  # 读取ini文件

result = config.sections()  # 获取节点
print(result)
for i in result:
    sheet = wb.create_sheet(i)  # 为每个节点创建一个新的sheet
    sheet['A1'] = '键'
    sheet['B1'] = '值'  # 改变内容,表头
    cell = sheet['A1':'B1']
    sheet['A1'].font = Font(name='黑体', color='FFFFFF')  # 改变字体颜色
    sheet['B1'].font = Font(name='黑体', color='FFFFFF')

    sheet['A1'].fill = PatternFill('solid', fgColor='0000FF')  # 填充背景颜色
    sheet['B1'].fill = PatternFill('solid', fgColor='0000FF')

    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)  # 设置对齐方式
    sheet['B1'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True)

    sheet['A1'].border = Border(
        top=Side(style="dashed", color="000000"),
        bottom=Side(style="dashed", color="000000"),
        left=Side(style="dashed", color="000000"),
        right=Side(style="dashed", color="000000"),
    )   #边框
    sheet['B1'].border = Border(
        top=Side(style="dashed", color="000000"),
        bottom=Side(style="dashed", color="000000"),
        left=Side(style="dashed", color="000000"),
        right=Side(style="dashed", color="000000"),
    )
    for row, (key, value) in enumerate(config.items(i), 2):  # 获取键，值，行数
        # print(row, key, value)
        sheet.cell(row, 1).value = key
        sheet.cell(row, 2).value = value

        sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                                                 wrap_text=True)
        sheet.cell(row, 2).alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                                                 wrap_text=True)
        sheet.cell(row, 1).border = Border(
            top=Side(style="dashed", color="000000"),
            bottom=Side(style="dashed", color="000000"),
            left=Side(style="dashed", color="000000"),
            right=Side(style="dashed", color="000000"),
        )
        sheet.cell(row, 2).border = Border(
            top=Side(style="dashed", color="000000"),
            bottom=Side(style="dashed", color="000000"),
            left=Side(style="dashed", color="000000"),
            right=Side(style="dashed", color="000000"),
        )

wb.save(os.path.join(path, 'new.xlsx'))
'''

# 4.
# 下载文件,将下载的文件保存到当前执行脚本同级目录下 /files/package/ 目录下（且文件名为HtmlStore.zip）
# 在将下载下来的文件解压到 /files/html/ 目录下
import requests
import shutil

path = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(path, 'files/package')
if not os.path.exists(file_path):  # 判断路径是否存在
    os.makedirs(file_path)  # 不存在就创建
file_url = 'https://files.cnblogs.com/files/wupeiqi/HtmlStore.zip'
res = requests.get(url=file_url)
zip_file_path = os.path.join(file_path, file_url.split('/')[-1])
with open(zip_file_path, mode='wb') as dowload:
    dowload.write(res.content)  # 下载文件
dowload_path = os.path.join(path, 'files', 'html')
shutil.unpack_archive(filename=zip_file_path, extract_dir=dowload_path, format='zip')
