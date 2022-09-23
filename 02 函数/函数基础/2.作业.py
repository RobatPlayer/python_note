# 1.请定义一个函数，用于计算一个字符串中字符a出现的次数，并通过return返回
def the_num_of(str):
    num = 0
    for i in str:
        if i == 'a':
            num += 1
    return num


result = the_num_of('zhangsan')
print(result)


# 2.写函数，判断用户传入的一个值（字符串或列表或元组任意）长度是否大于5，并返回真假。
def long(all_type):
    if len(all_type) > 5:
        return True
    return False


result = long([123, 456, 78, 4, 5, 3])
print(result)


# 3.写函数，接收两个数字参数，返回比较大的那个数字（等于时返回两个中的任意一个都可以）。
def big(a1, a2):
    if a1 > a2:
        return a1
    return a2


result = big(10, 20)
print(result)


# 4.写函数，函数接收四个参数分别是：姓名，性别，年龄，学历，然后将这四个值通过"*"拼接起来并追加到一个student_msg.txt文件中。
# 定义一个判断路径是否存在的函数
def file_exists():
    import os
    path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(path)
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    return file_path


def pocket(name, sex, age, education):
    result = '*'.join([name, sex, age, education])
    with open(f'{file_exists()}/student_msg.txt', mode='w', encoding='utf-8') as file:
        file.write(result)
        return 'over'


result = pocket('张三', '男', '21', '本科')
print(result)


# 5.补充代码，实现如下功能：
# 【位置1】读取文件中的每一行数据，将包含特定关键的数据筛选出来，并以列表的形式返回。
# 【位置1】文件不存在，则返回None
# 【位置2】文件不存在，输出"文件不存在"，否则循环输出匹配成功的每一行数据。
def select_content(file_path, key):
    import os
    if not os.path.exists(file_path):
        return
    l = []
    with open(file_path, mode='r', encoding='utf-8') as file:
        for line in file:
            if key in line:
                l.append(key)
                return l


result = select_content("files/xxx.txt", "股票")
if not result:
    print('文件不存在')
else:
    print(result)

# 6.补充代码，实现敏感词替换的功能。将字符串origin中中的敏感词替换为 **，最后将替换好的值返回。
'''
def change_string(origin):
    data_list = ["苍老师", "波多老师", "大桥"]
    for i in data_list:
        if i in origin:
            origin = origin.replace(i, '**')
    return origin


text = input("请输入内容：")
result = change_string(text)
print(result)
'''
# 7.基于函数实现用户认证，要求：写函数，读取的用户信息并构造为字典（用户信息存放在`files / user.xlsx`文件中）
# 构造的字典格式如下user_dict = {"用户名": "密码"}
# 用户输入用户名和密码，进行校验。（且密码都是密文，所以，需要将用户输入的密码进行加密，然后再与Excel中的密文密码进行比较）
import hashlib
import os
from openpyxl import workbook, load_workbook


def encrypt(origin):
    origin_bytes = origin.encode('utf-8')
    md5_object = hashlib.md5()
    md5_object.update(origin_bytes)
    return md5_object.hexdigest()


def file_exists():
    path = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(path, 'files')
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    return file_path


# 创建用户
dict = {}


def create(name, pwd):
    pwd = encrypt(pwd)
    dict.update({name: pwd})
    wb = workbook.Workbook()  # 创建一个Excel文件
    sheet = wb['Sheet']
    sheet.title = '加密数据'
    sheet.cell(1, 1).value = '用户名'
    sheet.cell(1, 2).value = '密码'  # 怎么将每次输入的内容加到Excel里面
    wb.save(f'{file_exists()}/user.xlsx')
    return dict


create('张三', '123456')
result = create('李四', '456789')
print(result)


# 验证用户
def register(name, pwd):
    pwd = encrypt(pwd)
    for key, value in result.items():
        if name == key and pwd == value:
            return '登录成功'

    else:
        return '账号或密码错误'


display1 = register('张三', '123456')
print(display1)
display2 = register('王五', '48456')
print(display2)
